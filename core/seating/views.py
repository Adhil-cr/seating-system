from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
from accounts.decorators import admin_required
from students.models import Student
from halls.models import Hall
from exams.models import Exam
from .models import SeatingAllocation, Seat
from .allocator_service import run_full_allocation_pipeline
import pandas as pd
from django.db import transaction

import json
import re
from io import BytesIO
from json import JSONDecodeError

# PDF
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils.text import slugify


# EXPORT: department abbreviations and colors
DEPT_ABBR = {
    "Automobile Engineering": "AE",
    "Mechanical Engineering": "ME",
    "Electrical & Electronics Engineering": "EEE",
    "Computer Engineering": "CT",
    "Electronics & Communication": "EC",
    "Civil Engineering": "CE",
}

# EXPORT: color scheme for departments
DEPT_COLORS = {
    "AE": {"bg": "#E3F2FD", "border": "#90CAF9", "text": "#0D47A1"},
    "ME": {"bg": "#FFF8E1", "border": "#FFCC80", "text": "#4E342E"},
    "EEE": {"bg": "#F3E5F5", "border": "#CE93D8", "text": "#4A148C"},
    "CT": {"bg": "#E8F5E9", "border": "#A5D6A7", "text": "#1B5E20"},
    "EC": {"bg": "#E0F7FA", "border": "#80DEEA", "text": "#006064"},
    "CE": {"bg": "#FCE4EC", "border": "#F48FB1", "text": "#880E4F"},
}

# EXPORT: fallback colors
DEPT_COLORS_FALLBACK = {"bg": "#F5F5F5", "border": "#BDBDBD", "text": "#424242"}


# EXPORT: helpers
def get_dept_abbr(dept_name):
    if not dept_name:
        return ""
    if dept_name in DEPT_ABBR:
        return DEPT_ABBR[dept_name]
    return "".join(word[0].upper() for word in dept_name.split())[:5]


def get_dept_colors(abbr):
    return DEPT_COLORS.get(abbr, DEPT_COLORS_FALLBACK)


# EXPORT: safe filenames for downloads
def _export_filename(exam, ext):
    name = slugify(exam.name) or "exam"
    session = slugify(str(exam.session)) or "session"
    date = str(exam.date)
    return f"seating_{name}_{date}_{session}.{ext}"


def _export_branch_filename(exam):
    name = slugify(exam.name) or "exam"
    session = slugify(str(exam.session)) or "session"
    date = str(exam.date)
    return f"branchwise_{name}_{date}_{session}.xlsx"


# -------------------------
# PREVIEW SEATING
# -------------------------

@admin_required
def preview_seating(request):
    total_students = Student.objects.filter(user=request.user).count()
    halls = Hall.objects.filter(user=request.user)

    total_capacity = sum(hall.capacity for hall in halls)

    return JsonResponse({
        "students": total_students,
        "capacity": total_capacity,
        "can_generate": total_capacity >= total_students
    })


# -------------------------
# GENERATE SEATING
# -------------------------

@admin_required
@transaction.atomic
def generate_seating(request):

    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    exam_id = data.get("exam_id")
    selected_halls = data.get("selected_halls", [])

    if not exam_id:
        return JsonResponse({"error": "exam_id required"}, status=400)

    exam = Exam.objects.filter(id=exam_id, user=request.user).first()
    if not exam:
        return JsonResponse({"error": "Exam not found"}, status=404)

    # Prevent duplicate generation (ignore stale allocations with zero seats)
    existing_allocation = SeatingAllocation.objects.filter(exam=exam).order_by("-id").first()
    if existing_allocation:
        if Seat.objects.filter(allocation=existing_allocation).exists():
            return JsonResponse({"error": "Seating already generated for this exam"}, status=400)
        # Stale allocation from a failed run – clean it up
        existing_allocation.delete()

    if selected_halls:
        halls_qs = Hall.objects.filter(id__in=selected_halls, user=request.user)
    else:
        halls_qs = Hall.objects.filter(user=request.user)

    halls = list(halls_qs)
    if not halls:
        return JsonResponse({"error": "No halls selected"}, status=400)

    subject_codes = list(getattr(exam, "subject_codes", []) or [])
    if not subject_codes:
        subject_codes = list(exam.subjects.values_list("code", flat=True))
    subject_codes = [str(code).strip().replace(".0", "") for code in subject_codes if str(code).strip()]

    total_students = Student.objects.filter(
        user=request.user,
        subjects__code__in=subject_codes
    ).distinct().count()
    total_capacity = sum(h.capacity for h in halls)

    if total_capacity < total_students:
        return JsonResponse({"error": "Not enough seats available"}, status=400)

    # -----------------------------
    # STEP 1: Run full algorithm pipeline
    # -----------------------------
    max_subject_per_hall = data.get("max_subject_per_hall")
    relax_subject_limit = data.get("relax_subject_limit")
    if relax_subject_limit:
        max_subject_per_hall = total_students
    elif max_subject_per_hall is not None:
        try:
            max_subject_per_hall = int(max_subject_per_hall)
        except (TypeError, ValueError):
            return JsonResponse({"error": "max_subject_per_hall must be an integer"}, status=400)
        if max_subject_per_hall <= 0:
            return JsonResponse({"error": "max_subject_per_hall must be positive"}, status=400)

    try:
        csv_output = run_full_allocation_pipeline(
            exam,
            halls=halls_qs,
            max_subject_per_hall=max_subject_per_hall
        )
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as e:
        import traceback
        print("\n\n===== ERROR START =====")
        print(str(e))
        traceback.print_exc()
        print("===== ERROR END =====\n\n")

        return JsonResponse({
            "error": str(e)
        }, status=500)

    # -----------------------------
    # STEP 2: Read algorithm output
    # -----------------------------
    df = pd.read_csv(csv_output)

    halls = list(halls_qs)

    if df.empty:
        return JsonResponse({"error": "Seating generation produced no rows"}, status=400)

    # Create allocation record (only after we know we have data)
    allocation = SeatingAllocation.objects.create(exam=exam)

    # -----------------------------
    # STEP 3: Convert to Django seats
    # -----------------------------
    for _, row in df.iterrows():

        hall_index = int(row["hall_id"]) - 1
        hall = halls[hall_index]

        seat_number = int(row["seat_number"])

        column_count = hall.columns

        row_no = (seat_number - 1) // column_count + 1
        column_no = (seat_number - 1) % column_count + 1

        student = Student.objects.filter(
            user=request.user,
            register_no=row["register_no"]
        ).first()

        if not student:
            raise ValueError(f"Student not found: {row['register_no']}")

        Seat.objects.create(
            allocation=allocation,
            hall=hall,
            row=row_no,
            column=column_no,
            student=student
        )

    return JsonResponse({"status": "seating generated using algorithm"})

# -------------------------
# VIEW SEATING
# -------------------------

@admin_required
def view_seating(request):

    exam_id = request.GET.get("exam_id")

    if not exam_id:
        return JsonResponse({"error": "exam_id required"}, status=400)

    exam = Exam.objects.filter(id=exam_id, user=request.user).first()
    if not exam:
        return JsonResponse({"error": "Exam not found"}, status=404)

    allocation = SeatingAllocation.objects.filter(
        exam_id=exam_id,
        exam__user=request.user
    ).last()

    if not allocation:
        return JsonResponse({"error": "No seating generated"}, status=404)

    seats = Seat.objects.filter(allocation=allocation).select_related("hall", "student").prefetch_related("student__subjects")

    # Department filter
    department_filter = request.GET.get("department")
    if department_filter:
        seats = seats.filter(student__department=department_filter)

    # Subject filter
    subject_filter = request.GET.get("subject")
    if subject_filter:
        seats = seats.filter(student__subjects__code=subject_filter)

    seats = seats.distinct()

    # FIX: deduplicate seats for multi-subject students
    seen_registers = set()
    unique_seats = []
    for seat in seats:
        reg_no = seat.student.register_no
        if reg_no in seen_registers:
            continue
        seen_registers.add(reg_no)
        unique_seats.append(seat)

    seats = unique_seats

    subject_codes = list(getattr(exam, "subject_codes", []) or [])
    if not subject_codes:
        subject_codes = list(exam.subjects.values_list("code", flat=True))
    subject_codes = [str(code).strip().replace(".0", "") for code in subject_codes if str(code).strip()]
    subject_code_set = set(subject_codes)

    halls_data = {}

    for seat in seats:

        hall_name = seat.hall.name

        if hall_name not in halls_data:
            halls_data[hall_name] = {
                "rows": seat.hall.rows,
                "columns": seat.hall.columns,
                "seats_per_bench": seat.hall.seats_per_bench,
                "seats": []
            }

        subjects = list(seat.student.subjects.values_list("code", flat=True))
        if subject_code_set:
            subjects = [code for code in subjects if code in subject_code_set]
        if not subjects and seat.student.subject_code:
            subjects = [str(seat.student.subject_code)]
        subjects = [str(code).strip().replace(".0", "") for code in subjects if str(code).strip()]
        subjects_str = ", ".join(sorted(set(subjects)))

        halls_data[hall_name]["seats"].append({
            "seat_number": (seat.row - 1) * seat.hall.columns + seat.column,
            "row": seat.row,
            "col": seat.column,
            "register_no": seat.student.register_no,
            "student_name": seat.student.name,
            "department": seat.student.department,
            "semester": seat.student.semester,
            "subjects": subjects_str,
            "is_multi_subject": len(set(subjects)) > 1
        })

    return JsonResponse({
        "exam": {
            "id": exam.id,
            "name": exam.name,
            "date": str(exam.date),
            "session": exam.session,
            "subject_codes": subject_codes,
        },
        "halls": [
            {
                "hall_name": hall_name,
                "room": "",
                "rows": halls_data[hall_name]["rows"],
                "cols": halls_data[hall_name]["columns"],
                "seats_per_bench": halls_data[hall_name].get("seats_per_bench", 1),
                "total_seats": halls_data[hall_name]["rows"] * halls_data[hall_name]["columns"],
                "seats": halls_data[hall_name]["seats"],
            }
            for hall_name in sorted(halls_data.keys())
        ]
    })


# -------------------------
# EXPORT PDF
# -------------------------

@admin_required
def export_pdf(request):
    # EXPORT: request validation
    exam_id = request.GET.get("exam_id")
    if not exam_id:
        return JsonResponse({"error": "exam_id required"}, status=400)

    # EXPORT: load exam and seats
    exam = get_object_or_404(Exam, id=exam_id, user=request.user)
    seats = (
        Seat.objects.filter(allocation__exam=exam)
        .select_related("hall", "student")
        .prefetch_related("student__subjects")
        .order_by("hall__name", "row", "column")
    )
    if not seats.exists():
        return JsonResponse({"error": "No seating generated"}, status=404)

    # EXPORT: hall metadata
    hall_meta = {}
    for seat in seats:
        hall_name = seat.hall.name
        if hall_name not in hall_meta:
            hall_meta[hall_name] = {
                "rows": seat.hall.rows,
                "cols": seat.hall.columns,
                "seats_per_bench": seat.hall.seats_per_bench,
            }

    # EXPORT: hall metadata
    hall_meta = {}
    for seat in seats:
        hall_name = seat.hall.name
        if hall_name not in hall_meta:
            hall_meta[hall_name] = {
                "rows": seat.hall.rows,
                "cols": seat.hall.columns,
                "seats_per_bench": seat.hall.seats_per_bench,
            }

    # EXPORT: subject codes list
    subject_codes = list(getattr(exam, "subject_codes", []) or [])
    if not subject_codes:
        subject_codes = list(exam.subjects.values_list("code", flat=True))
    subject_codes = [str(code).strip().replace(".0", "") for code in subject_codes if str(code).strip()]

    # EXPORT: deduplicate seats and collect subjects
    seen = {}
    for seat in seats:
        reg = seat.student.register_no
        if reg not in seen:
            subj_qs = seat.student.subjects.all()
            if subject_codes:
                subj_qs = subj_qs.filter(code__in=subject_codes)
            subj_list = list(subj_qs.values_list("code", flat=True))
            if not subj_list and seat.student.subject_code:
                subj_list = [str(seat.student.subject_code)]
            seat_no = (seat.row - 1) * seat.hall.columns + seat.column
            seen[reg] = {
                "hall": seat.hall.name,
                "seat_no": seat_no,
                "row": seat.row,
                "col": seat.column,
                "register_no": reg,
                "student_name": seat.student.name,
                "department": seat.student.department,
                "dept_abbr": get_dept_abbr(seat.student.department),
                "subjects": [str(code) for code in subj_list] if subj_list else [],
            }
        else:
            if getattr(seat, "subject_code", None):
                seen[reg]["subjects"].append(str(seat.subject_code))

    students = list(seen.values())
    for student in students:
        student["subjects"] = ", ".join(sorted(set(student["subjects"])))

    # EXPORT: group by hall
    from collections import defaultdict
    halls_data = defaultdict(list)
    for student in students:
        halls_data[student["hall"]].append(student)
    for hall_name in halls_data:
        halls_data[hall_name].sort(key=lambda x: x["seat_no"])
    sorted_halls = sorted(halls_data.keys())

    total_students = len(students)

    # EXPORT: PDF document setup
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    elements = []
    usable_width = doc.width

    # EXPORT: cover/title banner
    title_table = Table(
        [
            [exam.name],
            [f"Date: {exam.date}  |  Session: {exam.session}  |  Total: {total_students} students"],
            [f"Subjects: {', '.join(subject_codes)}"],
        ],
        colWidths=[usable_width],
    )
    title_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("TEXTCOLOR", (0, 1), (-1, 2), colors.HexColor("#B3C6E0")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 16),
        ("FONTNAME", (0, 1), (-1, 2), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, 1), 9),
        ("FONTSIZE", (0, 2), (-1, 2), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 1), (-1, 2), 6),
        ("BOTTOMPADDING", (0, 1), (-1, 2), 6),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 6 * mm))

    # EXPORT: department legend
    legend_abbrs = ["AE", "ME", "EEE", "CT", "EC", "CE"]
    legend_table = Table([legend_abbrs], colWidths=[usable_width / 6] * 6)
    legend_style = [("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD"))]
    for idx, abbr in enumerate(legend_abbrs):
        colors_map = get_dept_colors(abbr)
        legend_style.append(("BACKGROUND", (idx, 0), (idx, 0), colors.HexColor(colors_map["bg"])))
        legend_style.append(("TEXTCOLOR", (idx, 0), (idx, 0), colors.HexColor(colors_map["text"])))
    legend_table.setStyle(TableStyle(legend_style))
    elements.append(legend_table)
    elements.append(Spacer(1, 3 * mm))

    # EXPORT: summary heading
    summary_label = Table([["Hall Summary"]], colWidths=[usable_width])
    summary_label.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1F3864")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(summary_label)
    elements.append(Spacer(1, 2 * mm))

    # EXPORT: summary table data
    dept_order = ["AE", "ME", "EEE", "CT", "EC", "CE"]
    summary_rows = [["Hall", "Seats/Bench", "Students", "AE", "ME", "EEE", "CT", "EC", "CE", "Subjects"]]
    total_counts = {abbr: 0 for abbr in dept_order}
    total_students_sum = 0

    for hall_name in sorted_halls:
        hall_students = halls_data[hall_name]
        hall_count = len(hall_students)
        total_students_sum += hall_count
        dept_counts = {abbr: 0 for abbr in dept_order}
        subjects_set = set()
        for s in hall_students:
            if s["dept_abbr"] in dept_counts:
                dept_counts[s["dept_abbr"]] += 1
            for code in s["subjects"].split(","):
                code = code.strip()
                if code:
                    subjects_set.add(code)
        for abbr in dept_order:
            total_counts[abbr] += dept_counts[abbr]
        seats_per_bench = hall_meta.get(hall_name, {}).get("seats_per_bench", 1)
        summary_rows.append([
            hall_name,
            seats_per_bench,
            hall_count,
            dept_counts["AE"],
            dept_counts["ME"],
            dept_counts["EEE"],
            dept_counts["CT"],
            dept_counts["EC"],
            dept_counts["CE"],
            ", ".join(sorted(subjects_set)),
        ])

    summary_rows.append([
        "TOTAL",
        "",
        total_students_sum,
        total_counts["AE"],
        total_counts["ME"],
        total_counts["EEE"],
        total_counts["CT"],
        total_counts["EC"],
        total_counts["CE"],
        "",
    ])

    summary_col_widths = [
        12 * mm,  # Hall
        20 * mm,  # Seats/Bench
        14 * mm,  # Students
        13 * mm,
        13 * mm,
        13 * mm,
        13 * mm,
        13 * mm,
        13 * mm,
        usable_width - (12 + 20 + 14 + 13 * 6) * mm,
    ]
    summary_table = Table(summary_rows, colWidths=summary_col_widths)
    summary_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DEDEDE")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
        ("FONTSIZE", (0, 1), (-1, -2), 8),
        ("ALIGN", (0, 1), (-2, -2), "CENTER"),
        ("ALIGN", (-1, 1), (-1, -2), "LEFT"),
        ("FONTSIZE", (-1, 1), (-1, -2), 7),
        ("TEXTCOLOR", (-1, 1), (-1, -2), colors.HexColor("#595959")),
    ]
    for idx in range(1, len(summary_rows) - 1):
        bg = "#F5F8FC" if idx % 2 == 0 else "#FFFFFF"
        summary_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor(bg)))
    total_row_idx = len(summary_rows) - 1
    summary_style.append(("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), colors.HexColor("#E8EFF7")))
    summary_style.append(("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"))
    summary_style.append(("FONTSIZE", (0, total_row_idx), (-1, total_row_idx), 8))
    summary_table.setStyle(TableStyle(summary_style))

    elements.append(summary_table)
    elements.append(PageBreak())

    # EXPORT: hall pages
    for hall_idx, hall_name in enumerate(sorted_halls):
        hall_students = halls_data[hall_name]
        hall_count = len(hall_students)
        dept_counts = {}
        for s in hall_students:
            dept_counts[s["dept_abbr"]] = dept_counts.get(s["dept_abbr"], 0) + 1
        active_depts = [abbr for abbr in dept_order if dept_counts.get(abbr)]
        other_depts = sorted([abbr for abbr in dept_counts if abbr not in active_depts])
        active_depts.extend(other_depts)

        seats_per_bench = hall_meta.get(hall_name, {}).get("seats_per_bench", 1)
        header_table = Table(
            [[
                f"HALL  {hall_name}",
                f"{exam.date}  ·  {exam.session}  ·  {hall_count} students  ·  {seats_per_bench} seats/bench"
            ]],
            colWidths=[50 * mm, usable_width - 50 * mm],
        )
        header_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1F3864")),
            ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
            ("TEXTCOLOR", (1, 0), (1, 0), colors.HexColor("#B3C6E0")),
            ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (0, 0), 14),
            ("FONTNAME", (1, 0), (1, 0), "Helvetica"),
            ("FONTSIZE", (1, 0), (1, 0), 8),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (0, 0), 10),
            ("RIGHTPADDING", (1, 0), (1, 0), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(header_table)

        breakdown_text = "Dept breakdown:  " + "   |   ".join(
            f"{abbr}: {dept_counts[abbr]}" for abbr in active_depts
        )
        breakdown_table = Table([[breakdown_text]], colWidths=[usable_width])
        breakdown_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F2F2")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#595959")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(breakdown_table)

        if active_depts:
            key_table = Table([active_depts], colWidths=[usable_width / len(active_depts)] * len(active_depts))
            key_style = [("ALIGN", (0, 0), (-1, -1), "CENTER"),
                         ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                         ("FONTSIZE", (0, 0), (-1, -1), 7),
                         ("TOPPADDING", (0, 0), (-1, -1), 3),
                         ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                         ("BOX", (0, 0), (-1, -1), 0.3, colors.HexColor("#BDBDBD")),
                         ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BDBDBD"))]
            for idx, abbr in enumerate(active_depts):
                c_map = get_dept_colors(abbr)
                key_style.append(("BACKGROUND", (idx, 0), (idx, 0), colors.HexColor(c_map["bg"])))
                key_style.append(("TEXTCOLOR", (idx, 0), (idx, 0), colors.HexColor(c_map["text"])))
            key_table.setStyle(TableStyle(key_style))
            elements.append(key_table)

        elements.append(Spacer(1, 1 * mm))

        table_data = [[
            "Seat",
            "Row",
            "Col",
            "Register No",
            "Student Name",
            "Dept",
            "Subjects",
            "Signature",
        ]]
        for s in hall_students:
            table_data.append([
                s["seat_no"],
                s["row"],
                s["col"],
                s["register_no"],
                s["student_name"],
                s["dept_abbr"],
                s["subjects"],
                "",
            ])

        col_widths = [
            10 * mm,
            8 * mm,
            8 * mm,
            22 * mm,
            82 * mm,
            10 * mm,
            18 * mm,
            28 * mm,
        ]
        hall_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        hall_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DEDEDE")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
            ("LINEAFTER", (6, 0), (6, -1), 0.5, colors.HexColor("#BDBDBD")),
        ]
        for row_idx in range(1, len(table_data)):
            abbr = table_data[row_idx][5]
            c_map = get_dept_colors(abbr)
            hall_style.append(("BACKGROUND", (0, row_idx), (6, row_idx), colors.HexColor(c_map["bg"])))
            hall_style.append(("BACKGROUND", (7, row_idx), (7, row_idx), colors.white))
            hall_style.append(("TEXTCOLOR", (5, row_idx), (5, row_idx), colors.HexColor(c_map["text"])))
            hall_style.append(("FONTNAME", (0, row_idx), (0, row_idx), "Helvetica-Bold"))
            hall_style.append(("FONTSIZE", (0, row_idx), (7, row_idx), 8))
            hall_style.append(("FONTSIZE", (6, row_idx), (6, row_idx), 7))
            hall_style.append(("TEXTCOLOR", (6, row_idx), (6, row_idx), colors.HexColor("#595959")))
            hall_style.append(("FONTNAME", (5, row_idx), (5, row_idx), "Helvetica-Bold"))
            hall_style.append(("FONTSIZE", (5, row_idx), (5, row_idx), 7))
            hall_style.append(("ALIGN", (0, row_idx), (3, row_idx), "CENTER"))
            hall_style.append(("ALIGN", (5, row_idx), (7, row_idx), "CENTER"))
            hall_style.append(("ALIGN", (4, row_idx), (4, row_idx), "LEFT"))
            hall_style.append(("LEFTPADDING", (4, row_idx), (4, row_idx), 4))
        hall_table.setStyle(TableStyle(hall_style))
        elements.append(hall_table)

        elements.append(Spacer(1, 2 * mm))
        footer_table = Table(
            [[
                f"Hall {hall_name}  ·  {hall_count} students  ·  {exam.date} {exam.session}",
                "Invigilator: ____________________________",
            ]],
            colWidths=[usable_width * 0.6, usable_width * 0.4],
        )
        footer_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F2F2")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#595959")),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (0, 0), 8),
            ("RIGHTPADDING", (1, 0), (1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BOX", (0, 0), (-1, -1), 0.3, colors.HexColor("#BDBDBD")),
        ]))
        elements.append(footer_table)

        if hall_idx < len(sorted_halls) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="{_export_filename(exam, "pdf")}"'
    )
    return response


# -------------------------
# EXPORT EXCEL
# -------------------------

@admin_required
def export_excel(request):
    # EXPORT: request validation
    exam_id = request.GET.get("exam_id")
    if not exam_id:
        return JsonResponse({"error": "exam_id required"}, status=400)

    # EXPORT: load exam and seats
    exam = get_object_or_404(Exam, id=exam_id, user=request.user)
    seats = (
        Seat.objects.filter(allocation__exam=exam)
        .select_related("hall", "student")
        .prefetch_related("student__subjects")
        .order_by("hall__name", "row", "column")
    )
    if not seats.exists():
        return JsonResponse({"error": "No seating generated"}, status=404)

    # EXPORT: hall metadata
    hall_meta = {}
    for seat in seats:
        hall_name = seat.hall.name
        if hall_name not in hall_meta:
            hall_meta[hall_name] = {
                "rows": seat.hall.rows,
                "cols": seat.hall.columns,
                "seats_per_bench": seat.hall.seats_per_bench,
            }

    # EXPORT: subject codes list
    subject_codes = list(getattr(exam, "subject_codes", []) or [])
    if not subject_codes:
        subject_codes = list(exam.subjects.values_list("code", flat=True))
    subject_codes = [str(code).strip().replace(".0", "") for code in subject_codes if str(code).strip()]

    # EXPORT: deduplicate seats and collect subjects
    seen = {}
    for seat in seats:
        reg = seat.student.register_no
        if reg not in seen:
            subj_qs = seat.student.subjects.all()
            if subject_codes:
                subj_qs = subj_qs.filter(code__in=subject_codes)
            subj_list = list(subj_qs.values_list("code", flat=True))
            if not subj_list and seat.student.subject_code:
                subj_list = [str(seat.student.subject_code)]
            seat_no = (seat.row - 1) * seat.hall.columns + seat.column
            seen[reg] = {
                "hall": seat.hall.name,
                "seat_no": seat_no,
                "row": seat.row,
                "col": seat.column,
                "register_no": reg,
                "student_name": seat.student.name,
                "department": seat.student.department,
                "dept_abbr": get_dept_abbr(seat.student.department),
                "subjects": [str(code) for code in subj_list] if subj_list else [],
            }
        else:
            if getattr(seat, "subject_code", None):
                seen[reg]["subjects"].append(str(seat.subject_code))

    students = list(seen.values())
    for student in students:
        student["subjects"] = ", ".join(sorted(set(student["subjects"])))

    # EXPORT: group by hall
    from collections import defaultdict
    halls_data = defaultdict(list)
    for student in students:
        halls_data[student["hall"]].append(student)
    for hall_name in halls_data:
        halls_data[hall_name].sort(key=lambda x: x["seat_no"])
    sorted_halls = sorted(halls_data.keys())

    # EXPORT: workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    def _xl_color(value):
        value = value.replace("#", "").upper()
        if len(value) == 6:
            return f"FF{value}"
        return value

    thin_grey = Side(style="thin", color=_xl_color("#BDBDBD"))
    border_grey = Border(left=thin_grey, right=thin_grey, top=thin_grey, bottom=thin_grey)

    # EXPORT: summary header rows
    ws.merge_cells("A1:K1")
    ws["A1"] = f"{exam.name}  —  {exam.date}  |  Session: {exam.session}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=_xl_color("#1F3864"))
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Subjects: {', '.join(subject_codes)}"
    ws["A2"].font = Font(name="Arial", size=10, color=_xl_color("#1F3864"))
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = PatternFill("solid", fgColor=_xl_color("#D6E4F0"))
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:K3")
    ws["A3"] = (
        "Department colour key:   AE = Automobile   ME = Mechanical   "
        "EEE = Elec & Electronics   CT = Computer   EC = Electronics & Comm   CE = Civil"
    )
    ws["A3"].font = Font(name="Arial", size=9, color=_xl_color("#595959"))
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].fill = PatternFill("solid", fgColor=_xl_color("#F2F2F2"))
    ws.row_dimensions[3].height = 16

    # EXPORT: summary headers
    headers = ["Hall", "Seats/Bench", "Total Students", "AE", "ME", "EEE", "CT", "EC", "CE", "Subjects", "Session"]
    ws.append([])
    ws.append(headers)
    header_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_xl_color("#2E75B6"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_grey
    ws.row_dimensions[header_row].height = 22

    # EXPORT: summary data rows
    dept_order = ["AE", "ME", "EEE", "CT", "EC", "CE"]
    total_counts = {abbr: 0 for abbr in dept_order}
    total_students_sum = 0
    data_start = header_row + 1
    current_row = data_start

    for idx, hall_name in enumerate(sorted_halls):
        hall_students = halls_data[hall_name]
        hall_count = len(hall_students)
        total_students_sum += hall_count
        dept_counts = {abbr: 0 for abbr in dept_order}
        subjects_set = set()
        for s in hall_students:
            if s["dept_abbr"] in dept_counts:
                dept_counts[s["dept_abbr"]] += 1
            for code in s["subjects"].split(","):
                code = code.strip()
                if code:
                    subjects_set.add(code)
        for abbr in dept_order:
            total_counts[abbr] += dept_counts[abbr]

        seats_per_bench = hall_meta.get(hall_name, {}).get("seats_per_bench", 1)
        row_values = [
            hall_name,
            seats_per_bench,
            hall_count,
            dept_counts["AE"],
            dept_counts["ME"],
            dept_counts["EEE"],
            dept_counts["CT"],
            dept_counts["EC"],
            dept_counts["CE"],
            ", ".join(sorted(subjects_set)),
            exam.session,
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if col_idx == 10:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=10)
            cell.border = border_grey
        fill_color = "#EBF3FB" if idx % 2 == 1 else "#FFFFFF"
        for col_idx in range(1, 12):
            ws.cell(row=current_row, column=col_idx).fill = PatternFill("solid", fgColor=_xl_color(fill_color))
        current_row += 1

    # EXPORT: total row
    total_row = current_row
    total_values = [
        "TOTAL",
        "",
        total_students_sum,
        total_counts["AE"],
        total_counts["ME"],
        total_counts["EEE"],
        total_counts["CT"],
        total_counts["EC"],
        total_counts["CE"],
        "",
        "",
    ]
    for col_idx, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_xl_color("#1F3864"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_grey

    # EXPORT: summary column widths
    column_widths = [8, 12, 14, 8, 8, 8, 8, 8, 8, 35, 10]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # EXPORT: hall sheets
    for hall_name in sorted_halls:
        hall_students = halls_data[hall_name]
        hall_count = len(hall_students)
        dept_counts = {}
        for s in hall_students:
            dept_counts[s["dept_abbr"]] = dept_counts.get(s["dept_abbr"], 0) + 1
        dept_ordered = [abbr for abbr in dept_order if dept_counts.get(abbr)]
        dept_ordered.extend(sorted([abbr for abbr in dept_counts if abbr not in dept_ordered]))

        sheet = wb.create_sheet(title=f"Hall {hall_name}")
        seats_per_bench = hall_meta.get(hall_name, {}).get("seats_per_bench", 1)
        sheet.merge_cells("A1:H1")
        sheet["A1"] = (
            f"HALL  {hall_name}   ·   {exam.date}   ·   Session: {exam.session}   ·   "
            f"{hall_count} Students   ·   {seats_per_bench} seats/bench"
        )
        sheet["A1"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A1"].fill = PatternFill("solid", fgColor=_xl_color("#1F3864"))
        sheet.row_dimensions[1].height = 28

        sheet.merge_cells("A2:H2")
        breakdown_text = "  Dept breakdown:  " + "   |   ".join(
            f"{abbr}: {dept_counts[abbr]}" for abbr in dept_ordered
        )
        sheet["A2"] = breakdown_text
        sheet["A2"].font = Font(name="Arial", size=9, color=_xl_color("#595959"))
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A2"].fill = PatternFill("solid", fgColor=_xl_color("#F2F2F2"))
        sheet.row_dimensions[2].height = 16

        sheet.merge_cells("A3:H3")
        key_text = "  Colour key:  " + "   ".join(f"■ {abbr}" for abbr in dept_ordered)
        sheet["A3"] = key_text
        sheet["A3"].font = Font(name="Arial", size=9, color=_xl_color("#595959"))
        sheet["A3"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A3"].fill = PatternFill("solid", fgColor=_xl_color("#FAFAFA"))
        sheet.row_dimensions[3].height = 14

        hall_headers = ["Seat No", "Row", "Col", "Register No", "Student Name", "Dept", "Subjects", "Signature"]
        hall_header_row = 4
        for col_idx, header in enumerate(hall_headers, start=1):
            cell = sheet.cell(row=hall_header_row, column=col_idx, value=header)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_xl_color("#2E75B6"))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_grey
        sheet.row_dimensions[hall_header_row].height = 22

        data_row = hall_header_row + 1
        for student in hall_students:
            row_values = [
                student["seat_no"],
                student["row"],
                student["col"],
                student["register_no"],
                student["student_name"],
                student["dept_abbr"],
                student["subjects"],
                "",
            ]
            colors_map = get_dept_colors(student["dept_abbr"])
            row_fill = PatternFill("solid", fgColor=_xl_color(colors_map["bg"]))
            row_border = Border(
                left=Side(style="thin", color=_xl_color(colors_map["border"])),
                right=Side(style="thin", color=_xl_color(colors_map["border"])),
                top=Side(style="thin", color=_xl_color(colors_map["border"])),
                bottom=Side(style="thin", color=_xl_color(colors_map["border"])),
            )
            for col_idx, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=data_row, column=col_idx, value=value)
                if col_idx == 5:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.font = Font(name="Arial", size=10)
                elif col_idx == 6:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(
                        name="Arial", size=9, bold=True, color=_xl_color(colors_map["text"])
                    )
                elif col_idx == 7:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="Arial", size=9, color=_xl_color("#595959"))
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="Arial", size=10, bold=(col_idx == 1))

                if col_idx == 8:
                    cell.fill = PatternFill("solid", fgColor=_xl_color("#FFFFFF"))
                    cell.border = border_grey
                else:
                    cell.fill = row_fill
                    cell.border = row_border
            sheet.row_dimensions[data_row].height = 18
            data_row += 1

        footer_row = data_row
        sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
        sheet.merge_cells(start_row=footer_row, start_column=6, end_row=footer_row, end_column=8)
        left_cell = sheet.cell(row=footer_row, column=1, value=f"Total students in Hall {hall_name}: {hall_count}")
        right_cell = sheet.cell(row=footer_row, column=6, value="Invigilator Signature: ____________________________")
        for cell in (left_cell, right_cell):
            cell.font = Font(name="Arial", size=9, color=_xl_color("#595959"))
            cell.fill = PatternFill("solid", fgColor=_xl_color("#F2F2F2"))
            cell.border = border_grey
            cell.alignment = Alignment(
                horizontal="left" if cell == left_cell else "right",
                vertical="center"
            )
        sheet.row_dimensions[footer_row].height = 18

        hall_col_widths = [8, 6, 5, 14, 28, 8, 14, 20]
        for idx, width in enumerate(hall_col_widths, start=1):
            sheet.column_dimensions[chr(64 + idx)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_export_filename(exam, "xlsx")}"'
    )
    wb.save(response)
    return response


# -------------------------
# EXPORT BRANCH-WISE EXCEL
# -------------------------

@admin_required
def export_branchwise_excel(request):
    exam_id = request.GET.get("exam_id")
    if not exam_id:
        return JsonResponse({"error": "exam_id required"}, status=400)

    exam = get_object_or_404(Exam, id=exam_id, user=request.user)
    seats = (
        Seat.objects.filter(allocation__exam=exam)
        .select_related("hall", "student")
        .order_by("hall__name", "row", "column")
    )
    if not seats.exists():
        return JsonResponse({"error": "No seating generated"}, status=404)

    def _normalize_sheet_name(value):
        return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

    def _normalize_header(value):
        text = str(value or "").strip().lower()
        text = re.sub(r"[^a-z0-9]+", " ", text)
        return text.strip()

    def _find_header(ws):
        max_row = min(ws.max_row, 60)
        max_col = min(ws.max_column, 12)
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            col_map = {}
            for cell in row:
                text = _normalize_header(cell.value)
                if not text:
                    continue
                if "hall" in text and "hall" not in col_map:
                    col_map["hall"] = cell.column
                if "name" in text and "name" not in col_map:
                    col_map["name"] = cell.column
                if ("reg" in text or "register" in text or "prn" in text) and "reg" not in col_map:
                    col_map["reg"] = cell.column
                if ("sl" in text or "roll" in text) and "sl" not in col_map:
                    col_map["sl"] = cell.column
            if "name" in col_map and "hall" in col_map:
                return row[0].row, col_map
        return None, None

    def _set_cell_value(ws, row, col, value):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            return
        cell.value = value

    def _dept_to_branch(dept):
        if not dept:
            return ""
        d = dept.strip().lower()
        if "automobile" in d:
            return "AU"
        if "mechanical" in d:
            return "ME"
        if "civil" in d:
            return "CE"
        if "computer" in d:
            return "CT"
        if "electronics" in d and "communication" in d:
            return "EC"
        if "electrical" in d and "electronics" in d:
            return "EE"
        abbr = get_dept_abbr(dept)
        if abbr == "AE":
            return "AU"
        if abbr == "EEE":
            return "EE"
        return abbr

    def _normalize_semester(value):
        if value is None:
            return ""
        try:
            return int(value)
        except (TypeError, ValueError):
            match = re.search(r"\d+", str(value))
            return int(match.group()) if match else str(value).strip()

    def _reg_sort_key(value):
        text = str(value or "").strip()
        if not text:
            return (1, "")
        match = re.match(r"\d+", text)
        if match:
            return (0, int(match.group()), text)
        return (1, text.lower())

    students = {}
    for seat in seats:
        reg = str(seat.student.register_no).strip()
        if not reg:
            continue
        if reg not in students:
            students[reg] = {
                "register_no": reg,
                "name": seat.student.name,
                "department": seat.student.department,
                "semester": seat.student.semester,
                "hall": seat.hall.name,
            }

    if not students:
        return JsonResponse({"error": "No students found"}, status=404)

    template_path = settings.PROJECT_DIR / "static" / "samples" / "total_student_list_template.xlsx"
    if not template_path.exists():
        return JsonResponse({"error": "Template not found"}, status=500)

    wb = load_workbook(template_path)
    sheet_lookup = {_normalize_sheet_name(name): name for name in wb.sheetnames}

    def _branch_key_from_sheet(sheet_name):
        key = _normalize_sheet_name(sheet_name)
        match = re.search(r"s\\d+([a-z]+)$", key)
        if match:
            return match.group(1).upper()
        return ""

    def _update_semester_text(ws, new_sem):
        if not new_sem:
            return
        new_sem = str(new_sem)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    text = cell.value
                    if re.search(r"\\bsemester\\b", text, re.IGNORECASE):
                        text = re.sub(r"(semester\\s*:?\\s*)S?\\s*\\d+", r"\\1S " + new_sem, text, flags=re.IGNORECASE)
                    text = re.sub(r"\\bS\\s*\\d+\\b", "S " + new_sem, text)
                    text = re.sub(r"\\bS\\d+\\b", "S" + new_sem, text)
                    cell.value = text

    def _update_department_text(ws, dept_name):
        if not dept_name:
            return
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    text = cell.value
                    updated = text
                    if re.search(r"\\bdepartment\\b", updated, re.IGNORECASE):
                        if re.search(r"department\\s*of", updated, re.IGNORECASE):
                            updated = re.sub(
                                r"(department\\s*of\\s*).*",
                                r"\\1" + dept_name,
                                updated,
                                flags=re.IGNORECASE,
                            )
                        else:
                            updated = re.sub(
                                r"(department\\s*).*",
                                r"\\1" + dept_name,
                                updated,
                                flags=re.IGNORECASE,
                            )
                    if re.search(r"\\bprogramme\\b", updated, re.IGNORECASE):
                        updated = re.sub(
                            r"(programme\\s*:?\\s*).*",
                            r"\\1" + dept_name,
                            updated,
                            flags=re.IGNORECASE,
                        )
                    if updated != text:
                        cell.value = updated

    base_sheets = {}
    for name in wb.sheetnames:
        branch_key = _branch_key_from_sheet(name)
        if branch_key and branch_key not in base_sheets:
            base_sheets[branch_key] = name
    default_sheet_name = wb.sheetnames[0] if wb.sheetnames else None

    from collections import defaultdict
    groups = defaultdict(list)
    group_dept_name = {}
    for student in students.values():
        semester = _normalize_semester(student.get("semester"))
        branch = _dept_to_branch(student.get("department"))
        sheet_name = f"S{semester} {branch}".strip()
        groups[sheet_name].append(student)
        if sheet_name not in group_dept_name and student.get("department"):
            group_dept_name[sheet_name] = student.get("department")

    for sheet_key in list(groups.keys()):
        if _normalize_sheet_name(sheet_key) in sheet_lookup:
            continue
        match = re.search(r"s\\s*(\\d+)\\s*([a-z]+)", sheet_key, re.IGNORECASE)
        sem = match.group(1) if match else ""
        branch = match.group(2).upper() if match else ""
        base_sheet_name = base_sheets.get(branch) or default_sheet_name
        if not base_sheet_name:
            return JsonResponse({"error": "Template has no sheets to copy"}, status=500)
        new_sheet = wb.copy_worksheet(wb[base_sheet_name])
        new_sheet.title = sheet_key
        _update_semester_text(new_sheet, sem)
        _update_department_text(new_sheet, group_dept_name.get(sheet_key, ""))
        sheet_lookup[_normalize_sheet_name(sheet_key)] = sheet_key

    target_norms = {_normalize_sheet_name(name) for name in groups.keys()}
    unused_template_sheets = [
        name for name in wb.sheetnames
        if _normalize_sheet_name(name) not in target_norms
    ]
    for name in unused_template_sheets:
        if name in wb.sheetnames:
            del wb[name]

    sheets_to_delete = []
    for sheet_key, students_list in groups.items():
        sheet_name = sheet_lookup[_normalize_sheet_name(sheet_key)]
        ws = wb[sheet_name]
        header_row, col_map = _find_header(ws)
        if not header_row:
            return JsonResponse({
                "error": f"Header row not found in sheet: {sheet_name}"
            }, status=400)

        data_start = header_row + 1
        max_row = ws.max_row
        for row_idx in range(data_start, max_row + 1):
            for col_idx in col_map.values():
                _set_cell_value(ws, row_idx, col_idx, None)

        from collections import defaultdict
        halls = defaultdict(list)
        for student in students_list:
            halls[student["hall"]].append(student)

        current_row = data_start
        sl_no = 1
        for hall_name in sorted(halls.keys()):
            hall_students = sorted(halls[hall_name], key=lambda s: _reg_sort_key(s["register_no"]))
            for idx, student in enumerate(hall_students):
                if "sl" in col_map:
                    _set_cell_value(ws, current_row, col_map["sl"], sl_no)
                if "name" in col_map:
                    _set_cell_value(ws, current_row, col_map["name"], student["name"])
                if "reg" in col_map:
                    _set_cell_value(ws, current_row, col_map["reg"], student["register_no"])
                if "hall" in col_map:
                    hall_value = hall_name if idx == 0 else None
                    _set_cell_value(ws, current_row, col_map["hall"], hall_value)
                sl_no += 1
                current_row += 1

        if sl_no == 1:
            sheets_to_delete.append(sheet_name)

    for sheet_name in sheets_to_delete:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{_export_branch_filename(exam)}"'
    wb.save(response)
    return response
